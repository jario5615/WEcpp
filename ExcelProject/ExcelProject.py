from openpyxl import *



workbook = load_workbook("ExcelProjectBook.xlsx")
traineeSheet = workbook["ListOfTrainees"]
courseSheet = workbook["CourseDetails"]
trainerSheet = workbook["TrainerDetails"]
traineeCourseMap = workbook["MappingCourseTrainee"]
courseTrainerMap = workbook["MappingCourseTrainer"]

choice = ""

while choice != "0":
    trainees = [list(a) for a in traineeSheet.iter_rows(values_only=True)]
    courses = [list(a) for a in courseSheet.iter_rows(values_only=True)]
    trainers = [list(a) for a in trainerSheet.iter_rows(values_only=True)]

    print("Quit - 0")
    print("Add trainee - 1")
    print("Delete trainee - 2")
    print("Update trainee - 3")
    print("Add Course - 4")
    print("Delete Course - 5")
    print("Update Course - 6")
    print("Add trainer - 7")
    print("Delete trainer - 8")
    print("Update trainer - 9")
    print("Add trainer to course - 10")
    print("Remove trainer from course - 11")
    # Later mods

    choice = input("Enter numeric choice: ")
    match choice:
        case "0":
            break
        case "1":
            ID = input("Enter ID for new trainee: ")
            name = input("Enter name: ")
            course = input("Enter ID for course: ")
            background = input("Enter background: ")
            experience = input("Enter work experience: ")
            trainees.append([ID, name, course, background, experience])
            print("Trainee Added")
        case "2":
            ID = input("Enter ID for trainee to delete")
            for i in range(1, len(trainees)):
                if trainees[i][0] == ID:
                    trainees.pop(i)
                    print("Trainee Deleted")
                    break
            else:
                print("Trainee does not exist")
        case "3":
            ID = input("Enter ID for trainee to update")
            for i in range(1, len(trainees)):
                if trainees[i][0] == ID:
                    name = input("Enter name for trainee, or leave blank to pass: ")
                    course = input("Enter ID for course, or leave blank to pass: ")
                    background = input("Enter background, or leave blank to pass: ")
                    experience = input("Enter work experience, or leave blank to pass")
                    if name == "":
                        name = trainees[i][1]
                    if course == "":
                        course = trainees[i][2]
                    if background == "":
                        background = trainees[i][3]
                    if experience == "":
                        experience = trainees[i][4]
                    trainees[i] = [ID, name, course, background, experience]
                    print("Successfully updated info")
                    break
            else:
                print("Trainee does not exist")
        case "4":
            ID = input("Enter ID for new Course: ")
            desc = input("Enter description: ")
            courses.append([ID, desc])
            print("Course Added")
        case "5":
            ID = input("Enter ID for course to delete")
            for i in range(1, len(courses)):
                if courses[i][0] == ID:
                    courses.pop(i)
                    print("Course Deleted")
                    break
            else:
                print("Course does not exist")
        case "6":
            ID = input("Enter ID for course to update: ")
            for i in range(1, len(courses)):
                if courses[i][0] == ID:
                    desc = input("Enter new description: ")
                    courses[i] = [ID, desc]
                    print("Course updated")
                    break
            else:
                print("Course does not exist")
        case "7":
            ID = input("Enter ID for new trainer: ")
            name = input("Enter name: ")
            email = input("Enter email address: ")
            phone = input("Enter phone number: ")
            trainers.append([ID, name, email, phone,[]])
            print("Trainer Added")
        case "8":
            ID = input("Enter ID for trainer to delete: ")
            for i in range(1, len(trainers)):
                if trainers[i][0] == ID:
                    trainers.pop(i)
                    print("Trainer Deleted")
                    break
            else:
                print("Trainer does not exist")
        case "9":
            ID = input("Enter ID for trainer to update: ")
            for i in range(1, len(trainers)):
                if trainers[i][0] == ID:
                    name = input("Enter name, or leave blank to pass: ")
                    email = input("Enter email address, or leave blank to pass: ")
                    phone = input("Enter phone number, or leave blank to pass: ")
                    if name == "":
                        name = trainers[i][1]
                    if email == "":
                        email = trainers[i][2]
                    if phone == "":
                        phone = trainers[i][3]
                    trainers[i] = [ID, name, email, phone, trainers[i][4]]
                    print("Successfully updated info")
                    break
            else:
                print("Trainer does not exist")
        case "10":
            ID = input("Enter ID for trainer to add: ")
            for i in range(1, len(trainers)):
                if trainers[i][0] == ID:
                    course = input("Enter course to add trainer to: ")
                    for j in range(1, len(courses)):
                        if courses[j][0] == course:
                            trainers[i][4].append(course)
                            break
                    else:
                        print("Course does not exist")
                    break
            else:
                print("Trainer does not exist")
        case "11":
            ID = input("Enter ID for trainer to remove: ")
            for i in range(1, len(trainers)):
                if trainers[i][0] == ID:
                    course = input("Enter course to remove trainer from: ")
                    for j in range(1, len(trainers[i][4])):
                        if trainers[i][4][j] == course:
                            trainers[i][4].pop(j)
                            break
                    else:
                        print("Trainer not on course")
                    break
            else:
                print("Trainer does not exist")
        #Additional commands here


    for i in range(1, len(trainees)+1):
        traineeSheet["A" + str(i)] = trainees[i - 1][0]
        traineeSheet["B" + str(i)] = trainees[i - 1][1]
        traineeSheet["C" + str(i)] = trainees[i - 1][2]
        traineeSheet["D" + str(i)] = trainees[i - 1][3]
        traineeSheet["E" + str(i)] = trainees[i - 1][4]

        traineeCourseMap["A" + str(i)] = trainees[i - 1][0]
        traineeCourseMap["B" + str(i)] = trainees[i - 1][2]

    for i in range(1, len(courses)+1):
        courseSheet["A" + str(i)] = courses[i - 1][0]
        courseSheet["B" + str(i)] = courses[i - 1][1]

    k = 1

    for i in range(1, len(trainers)+1):
        trainerSheet["A" + str(i)] = trainers[i - 1][0]
        trainerSheet["B" + str(i)] = trainers[i - 1][1]
        trainerSheet["C" + str(i)] = trainers[i - 1][2]
        trainerSheet["D" + str(i)] = trainers[i - 1][3]
        '''
        for a in trainers[i - 1][4]:
            traineeCourseMap["A" + str(k)] = trainers[i - 1][0]
            traineeCourseMap["B" + str(k)] = a
            k += 1
        '''

    workbook.save(filename="ExcelProjectBook.xlsx")
